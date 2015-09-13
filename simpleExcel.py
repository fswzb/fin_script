# -*- coding:gbk -*-
import datetime
from openpyxl import Workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
ws['A2'] = datetime.datetime.now()

str = u'����'
ws['A3'] = str.encode('utf8')

str = '����'
s1 = str.decode('gbk')
ws['A4'] = s1.encode('utf8') 


strline = u'�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����'
strObj = strline.split(',')
print strObj
ws.append(strObj)

ws = wb.create_sheet()
ws.title = 'A2'

# Save the file
wb.save("z_excel.xlsx")
