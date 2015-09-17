# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

prepath = '../Data/'
allFileNum = 0
def loginfo():
	log = 1
	if (log==1):
		frame = None
		try:
			raise  ZeroDivisionError
		except  ZeroDivisionError:
			frame = sys.exc_info()[2].tb_frame.f_back
		print "%s in line %d" %(str(datetime.datetime.now()), frame.f_lineno)

def printPath(level, path,fileList):
	global allFileNum
	# �����ļ��У���һ���ֶ��Ǵ�Ŀ¼�ļ���
	dirList = []
	# �����ļ�
	#fileList = []
	# ����һ���б����а�����Ŀ¼��Ŀ������(google����)
	files = os.listdir(path)
	# �����Ŀ¼����
	dirList.append(str(level))
	for f in files:
		if(os.path.isdir(path + '/' + f)):
			# �ų������ļ��С���Ϊ�����ļ��й���
			if(f[0] == '.'):
				pass
			else:
				# ��ӷ������ļ���
				dirList.append(f)
		if(os.path.isfile(path + '/' + f)):
			# ����ļ�
			fileList.append(f)
	# ��һ����־ʹ�ã��ļ����б��һ�����𲻴�ӡ
	''' Ŀǰ����Ҫ���ļ���
	i_dl = 0
	for dl in dirList:
		if(i_dl == 0):
			i_dl = i_dl + 1
		else:
			# ��ӡ������̨�����ǵ�һ����Ŀ¼
			print '-' * (int(dirList[0])), dl
			# ��ӡĿ¼�µ������ļ��к��ļ���Ŀ¼����+1
			printPath((int(dirList[0]) + 1), path + '/' + dl)
	for fl in fileList:
		# ��ӡ�ļ�
		#print '-' * (int(dirList[0])), fl
		# ������һ���ж��ٸ��ļ�
		allFileNum = allFileNum + 1
		file = os.open(fl, "r")
	'''
	
	'''
	printPath(1, path, fileList)
	for fl in fileList:
		wkfile = path +"//"+ fl
		wb = load_workbook(wkfile)
		print "Worksheet range(s):", wb.get_named_ranges()
		print "Worksheet name(s):", wb.get_sheet_names()
		if "statistics" not in wb.get_sheet_names():
			print "NONNN"
			continue

		sheet_ranges_r = wb.get_sheet_by_name(name='statistics')
		print "YYYY:"
		print sheet_ranges_r
	'''	

# �����洢���ݵ��ֵ� 
data_dic = [] 

def parseFile(path, filename):
	wkfile = path +"/"+ filename
	print wkfile

	loginfo()
	wb = load_workbook(wkfile)
	#print "Worksheet range(s):", wb.get_named_ranges()
	#print "Worksheet name(s):", wb.get_sheet_names()
	loginfo()
	if "statistics" not in wb.get_sheet_names():
		return

	loginfo()
	ws = wb.get_sheet_by_name(name='statistics')
	loginfo()
	max_column = ws.max_column
	if max_column<7:
		print "column(%d) too short, please check" % max_column
		return

	print 111
	#�����ݴ浽�ֵ���
	pid = ''
	data_list = []
	for rx in range(1,ws.max_row):
		w1 = ws.cell(row = rx, column = 1).value
		w2 = ws.cell(row = rx, column = 2).value
		w3 = ws.cell(row = rx, column = 3).value
		w4 = ws.cell(row = rx, column = 4).value
		w5 = ws.cell(row = rx, column = 5).value
		w6 = ws.cell(row = rx, column = 6).value
		w7 = ws.cell(row = rx, column = 7).value
		if (w1 is None) and (w2 is None) and (w3 is None) and (w4 is None)\
			and (w5 is None) and (w6 is None) and (w7 is None):
			break
		if (w1 is None) or (w2 is None) or (w3 is None) or (w4 is None)\
			or (w5 is None) or (w6 is None) or (w7 is None):
			print "ĳ���¼ΪNone������ȷ", w1, w2, w3, w4, w5, w6, w7 
			continue
		temp_list = [w1,w2,w3,w4,w5,w6,w7]
		if rx==1:
			pid = w1
		else:
			data_list.append(temp_list)

	if cmp(pid, '')!=0:
		day_list = [pid, data_list]
		data_dic.append(day_list)

	
def updateFile(path, filename):
	fltlist = [200, 300, 600]
	wkfile = path +"/"+ filename

	dataObjLen = len(data_dic)
	if dataObjLen==0:
		print "û������"
		return
	
	wb = Workbook()
	ws = wb.active
	ws.title = 'statistics'
	
	ascid = 65
	row = 1
	for fltvol in fltlist:
		title = [fltvol, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg', ]
		number = len(title)
		for i in range(0,number):
			cell = chr(ascid+i) + str(row)
			ws[cell] = title[i]
		row += 1

		index = dataObjLen-1
		while index>=0:
			day_list = data_dic[index]
			index -= 1
			for data_list in day_list[1]:
				if data_list[0]!=fltvol:
					continue

				number = len(data_list)
				for i in range(0,number):
					cell = chr(ascid+i) + str(row)
					if i==0:
						ws[cell] = day_list[0]
					else:
						ws[cell] = data_list[i]
				row += 1
		#д����ͬ����Ȼ���һ��
		row += 1
	wb.save(wkfile)


	
if __name__ == '__main__':
	pindex = len(sys.argv)
	if pindex<2:
		sys.stderr.write("Usage: " +os.path.basename(sys.argv[0])+ " ����\n")
		exit(1);

	code = sys.argv[1]
	if (len(code) != 6):
		sys.stderr.write("Len should be 6\n")
		exit(1);

	head3 = code[0:3]
	result = (cmp(head3, "000")==0) or (cmp(head3, "002")==0) or (cmp(head3, "300")==0)
	if result is True:
		code = "sz" + code
	else:
		result = (cmp(head3, "600")==0) or (cmp(head3, "601")==0) or (cmp(head3, "603")==0)
		if result is True:
			code = "sh" + code
		else:
			print "�Ƿ�����:" +code+ "\n"
			exit(1);

	fileList = []
	path = prepath + code
	if not os.path.isdir(path):
		print "'"+ path +"' not a dir"
		exit(1)

	for (dirpath, dirnames, filenames) in os.walk(path):  
		#print('dirpath = ' + dirpath)
		i = 0
		for filename in filenames:
			extname = filename.split('.')[-1]
			if cmp(extname,"xlsx")!=0:
				continue
			prename = code+"_"
			if cmp(filename[0:9], prename[0:9])!=0:
				continue

			#print filename
			parseFile(path, filename)
			i += 1
			
		#�����õ����ļ��е��ļ����������ļ������ļ�
		break;

	if cmp('meg', '')!=0:
		stfile = '_'+ code +'__result.xlsx'
		updateFile(path, stfile)
		
