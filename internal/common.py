# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib
import urllib2
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook

#url = "http://vip.stock.finance.sina.com.cn/quotes_service/view/vMS_tradedetail.php?symbol=sz300001&date=2015-09-10&page=48"
#�ɽ�ʱ��	�ɽ���	�ǵ���	�۸�䶯	�ɽ���(��)		�ɽ���(Ԫ)	����
#	<tr ><th>11:29:48</th><td>14.57</td><td>-3.06%</td><td>+0.01</td><td>16</td><td>23,312</td><th><h1>������</h1></th></tr>
#<tr ><th>11:29:36</th><td>14.56</td><td>-3.13%</td><td>--</td><td>9</td><td>13,104</td><th><h6>����</h6></th></tr>
#<tr ><th>11:29:21</th><td>14.56</td><td>-3.13%</td><td>-0.02</td><td>10</td><td>14,560</td><th><h6>����</h6></th></tr>

class fitItem:
	volumn = 0
	buyvol = 0
	buyct = 0
	buyavg = 0
	sellvol = 0
	sellct = 0
	sellavg = 0
	def __init__(self, vol):
		self.volumn = vol
		self.buyvol = 0
		self.buyct = 0
		self.buyavg = 0
		self.sellvol = 0
		self.sellct = 0
		self.sellavg = 0

dftsarr = "0,200,300,600,900"
HandleMid = 1

def loginfo(flag=0):
	if (flag==1):
		frame = None
		try:
			raise  ZeroDivisionError
		except  ZeroDivisionError:
			frame = sys.exc_info()[2].tb_frame.f_back
		print "%s in line %d" %(str(datetime.datetime.now()), frame.f_lineno)

def parseCode(code):
	if (len(code) != 6):
		sys.stderr.write("Len should be 6\n")
		return (-1, '')

	head3 = code[0:3]
	result = (cmp(head3, "000")==0) or (cmp(head3, "002")==0) or (cmp(head3, "300")==0)
	if result is True:
		ncode = "sz" + code
	else:
		result = (cmp(head3, "600")==0) or (cmp(head3, "601")==0) or (cmp(head3, "603")==0)
		if result is True:
			ncode = "sh" + code
		else:
			print "�Ƿ�����:" +code+ "\n"
			return (-1, '')
	return (0, ncode)

def parseDate(qdate, today):
	dateObj = re.match(r'^(\d{4})(\d{2})(\d{2})', qdate)
	if (dateObj is None):
		dateObj = re.match(r'^(\d{2})(\d{2})', qdate)
		if (dateObj is None):
			print "�Ƿ����ڸ�ʽ��" +qdate+ ",������ʽ:YYYYMMDD or MMDD"
			return (-1, '')
		else:
			year = today.year
			month = int(dateObj.group(1))
			day = int(dateObj.group(2))
	else:
		year = int(dateObj.group(1))
		month = int(dateObj.group(2))
		day = int(dateObj.group(3))
	strdate = '%04d-%02d-%02d' %(year, month, day)
	return (0, strdate)

def parseTime(qtime, today):
	timeObj = re.match(r'^(\d{2}):(\d{2}):(\d{2})', qtime)
	if (timeObj is None):
		print "�Ƿ�ʱ���ʽ��" +qtime+ ",������ʽ:HH:MM:SS"
		return (-1, -1, -1, -1)

	hour = int(timeObj.group(1))
	minute = int(timeObj.group(2))
	second = int(timeObj.group(3))
	return (0, hour, minute, second)

#type:
#	1:Buy,	2:Sell
#����4����������Ϊ1�������۰봦��
def handle_volumn(exVolumn, dataObj, type, flag=0):
	dataObjLen = len(dataObj)
	chgVolumn = exVolumn
	if flag==1:
		chgVolumn = exVolumn/2

	if type==1:
		for j in range(0, dataObjLen):
			if exVolumn<dataObj[j].volumn:
				break;
			dataObj[j].buyvol += chgVolumn
			dataObj[j].buyct += 1
	elif type==2:
		for j in range(0, dataObjLen):
			if exVolumn<dataObj[j].volumn:
				break;
			dataObj[j].sellvol += chgVolumn
			dataObj[j].sellct += 1
			
def handle_middle_volumn(exVolumn, dataObj, exTime, fluctuate, increaseRange):
	if HandleMid!=1:
		return 0

	ret = 0
	dataObjLen = len(dataObj)

	#���⴦���һ�� (9:25) �ĳɽ�
	timeobj = exTime.split(':')
	hour = int(timeobj[0])
	minute = int(timeobj[1])
	if (hour==9 and (minute>24 and minute<30)):
		if increaseRange is None:
			print "______ It is NONE"
			return 0
		rangeobj = re.match(r'(.*)\%', increaseRange)
		if rangeobj is None:
			print "Not match string:", increaseRange
			return 0

		fltval = float(rangeobj.group(1))
		if fltval>0.001:
			handle_volumn(exVolumn, dataObj, 1)
		elif fltval<-0.001:
			handle_volumn(exVolumn, dataObj, 2)
		else:
			handle_volumn(exVolumn, dataObj, 1, 1)
			handle_volumn(exVolumn, dataObj, 2, 1)
		return 0

	if (fluctuate=='--'):
		handle_volumn(exVolumn, dataObj, 1, 1)
		handle_volumn(exVolumn, dataObj, 2, 1)
		return 0

	ftfluct = float(fluctuate)
	if (ftfluct>-0.021 and ftfluct<0.021):
		handle_volumn(exVolumn, dataObj, 1, 1)
		handle_volumn(exVolumn, dataObj, 2, 1)
	elif (ftfluct>1.0 or ftfluct<-1.0):
		print "??? Flucture is too big:", fluctuate
	elif (ftfluct>0.02):
		handle_volumn(exVolumn, dataObj, 1)
		ret = 1
	elif (ftfluct<-0.02):
		handle_volumn(exVolumn, dataObj, 2)
		ret = 2
	return ret

def write_statics(ws, fctime, dataObj, qdate):
	ws.title = 'statistics'

	ascid = 65
	row = 1
	if cmp(fctime, '')==0:
		title = [qdate, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg', ]
	else:
		title = [qdate, 'B', 'S', 'B_vol', 'S_vol', 'B_avg', 'S_avg', fctime]
	number = len(title)
	for i in range(0,number):
		cell = chr(ascid+i) + str(row)
		ws[cell] = title[i]

	dataObjLen = len(dataObj)
	for j in range(0, dataObjLen):
		list = []
		list.append(dataObj[j].volumn)
		
		buyvol = dataObj[j].buyvol
		buyct = dataObj[j].buyct
		
		sellvol = dataObj[j].sellvol
		sellct = dataObj[j].sellct
		
		list.append(buyvol)
		list.append(sellvol)
		list.append(buyct)
		list.append(sellct)
		if buyct==0:
			list.append(0)
		else:
			list.append(buyvol/buyct)
		if sellct==0:
			list.append(0)
		else:
			list.append(sellvol/sellct)
		list.append('')
		list.append(buyvol + sellvol)
		list.append(buyct + sellct)
		bsct = buyct + sellct
		if bsct==0:
			list.append(0)
		else:
			list.append((buyvol + sellvol)/bsct)
		
		row = row+1
		number = len(list)
		for i in range(0,number):
			cell = chr(ascid+i) + str(row)
			ws[cell] = list[i]

def handle_data(addcsv, prepath, bhist, url, code, qdate, sarr):
	if bhist==0:
		url = url +"?symbol="+ code
	elif bhist==1:
		url = url +"?date="+ qdate +"&symbol="+ code
	else:
		print "Unknown flag:", bhist
		return
	if HandleMid==0:
		print "Message: Ignore ������"

	if not os.path.isdir(prepath):
		os.makedirs(prepath)

	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	dataObjLen = len(dataObj)

	wb = Workbook()
	# grab the active worksheet
	ws = wb.active

	totalline = 0
	#���������ڲ�ͬ��ҳ�棬ͬʱ���ڣ������ظ�������Ҫ�����ظ����
	#��������ͬʱ�䣬��������ɽ�������Ҫ������
	lasttime = ''
	lastvol = 0
	pageFtime = ''
	bFtime = 0
	bFindHist = 0
	hisUrl = ''
	filename = code+ '_' + qdate
	fctime = ''
	if bhist==0:
		cur=datetime.datetime.now()
		fctime = '%02d:%02d' %(cur.hour, cur.minute)
		filename = '%s_%02d-%02d' %(filename, cur.hour, cur.minute)
	if addcsv==1:
		filecsv = prepath + filename + '.csv'
		fcsv = open(filecsv, 'w')
		strline = '�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����'
		fcsv.write(strline)
		fcsv.write("\n")

	strline = u'�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����,���̼�,�ǵ���,ǰ�ռ�,���̼�,��߼�,��ͼ�,�ɽ���,�ɽ���'
	strObj = strline.split(u',')
	ws.append(strObj)
	dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(����|����|������)\D')
	frameRe = re.compile(r'.*name=\"list_frame\" src=\"(.*)\" frameborder')
	keyw = '���̼�|�ǵ���|ǰ�ռ�|���̼�|��߼�|��ͼ�|�ɽ���|�ɽ���'
	infoRe = re.compile(r'\D+('+keyw+').*>(\+?-?\d+\.\d+)')
	excecount = 0
	stockInfo = []
	reloadUrl = 0
	noDataFlag = 0
	noDataKey = "�ù�Ʊû�н�������"
	i = 1

	for j in range(1,1000):
		urlall = url + "&page=" +str(i)
		#print "%d, %s" %(i,urlall)

		if excecount>10:
			print "Quit with exception i=", i
			break

		#����url���ӣ���ȡÿһҳ������
		req = urllib2.Request(urlall)
		try:
			res_data = urllib2.urlopen(req)
		except:
			print "Get URL except"
			excecount += 1
			continue
		else:
			excecount = 0
			pass

		flag = 0
		count = 0
		bFtime = 0
		
		#��ʼ��ȡÿһҳ���ص����ݣ����Ȳ���'�ɽ�ʱ��'/'���̼�'�����˴�������Ҫ������
		line = res_data.readline()
		if bhist==0:
			checkStr = '�ɽ�ʱ��'
		else:
			checkStr = '���̼�'
		#print "==========0000", line

		#���ҵ�'�ɽ�ʱ��'/'���̼�'�����²�������Ϊ'<script type='
		while line:
			index = line.find(checkStr)
			if (index<0):
				line = res_data.readline()
				continue
			else:
				checkStr = '<script type='
				break;

		#�ӵ�һҳ��ȡ�����̼�,�ǵ���,ǰ�ռ۵�����
		if bhist==1 and i==1:
			infoCount = 0
			while line:
				infoObj = infoRe.match(line)
				if infoObj:
					stockInfo.append(float(infoObj.group(2)))
				else:
					if infoCount<8:
						print "Parse fail", line
					break;
				infoCount += 1
				line = res_data.readline()
			#print stockInfo

		while line:
			#print line
			index = line.find(checkStr)
			if (index>=0):
				#�ҵ��ؼ���'<script type='���ҵ����������ݲ��÷�����׼����ȡ��һҳ��������
				#print "%s found, QUIT line='%s' "%(checkStr, line)
				break;

			#key = re.match(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(\+?-?\d+.\d+%)\D+(--|\+\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(����|����|������)\D', line)
			key = dtlRe.match(line)
			if key:
				#print key.groups(), sys._getframe().f_lineno 
				curtime = key.group(1)
				curvol = int(key.group(5))
				#��ס��ǰҳ��һ����ʱ��
				if (bFtime==0):
					timeobj = re.search(curtime, pageFtime)
					if timeobj:
						print "bFtime 0 quit"
						break
					pageFtime = curtime
					bFtime = 1

				timeobj = re.search(curtime, lasttime)
				if (timeobj and curvol==lastvol):
					count += 1
					line = res_data.readline()
					continue

				if (key.group(2)=="0.00") or (key.group(3)=="-100.00%"):
					print "page(%d) Price(%s) or range(%s) is invalid value"%(i, key.group(2), key.group(3))
					pass
				else:
					curprice = key.group(2)
					fluctuate = key.group(4)
					lasttime = curtime
					lastvol = curvol
					amount = key.group(6)
					obj = amount.split(',')
					amount = ''.join(obj)

					volume = int(key.group(5))
					updatestate = key.group(7)
					state = key.group(7)
					if cmp(state, '����')==0:
						handle_volumn(volume, dataObj, 2)
					elif cmp(state, '����')==0:
						handle_volumn(volume, dataObj, 1)
					elif cmp(state, '������')==0:
						ret = handle_middle_volumn(volume, dataObj, curtime, fluctuate, key.group(3))
						if ret==1:
							state = '����'
						elif ret==2:
							state = '����'

					if addcsv==1:
						strline = curtime +","+ key.group(2) +","+ key.group(3) +","+ key.group(4) +","+ key.group(5) +","+ amount +","+ key.group(7) + "\n"
						fcsv.write(strline)

					totalline += 1
					row = totalline+1
					cell = 'A' + str(row)
					ws[cell] = curtime
					cell = 'B' + str(row)
					ws[cell] = float(key.group(2))
					cell = 'C' + str(row)
					ws[cell] = key.group(3)
					cell = 'D' + str(row)
					if (fluctuate=='--'):
						ws[cell] = key.group(4)
					else:
						ftfluct = float(fluctuate)
						ws[cell] = ftfluct
					cell = 'E' + str(row)
					ws[cell] = int(key.group(5))
					cell = 'F' + str(row)
					ws[cell] = int(amount)
					cell = 'G' + str(row)
					s1 = state.decode('gbk')
					ws[cell] = s1
					
					if (row==2 and bhist==1):
						ascid = 72
						number = len(stockInfo)
						for j in range(0,number):
							cell = chr(ascid+j) + str(row)
							ws[cell] = stockInfo[j]

				count += 1
				line = res_data.readline()
				continue
			else:
				index = line.find(noDataKey)
				if (index>=0):
					#�ҵ��ؼ��֣���ǰ���Ժ��ҳ�涼û������
					#print "KEY word '%s' found, QUIT line='%s' "%(noDataKey, line)
					noDataFlag = 1
					break;

			if bhist==1:
				key = frameRe.match(line)
				if key:
					bFindHist = 1
					hisUrl = key.group(1)
					print key.groups()
					break
			
			line = res_data.readline()

		loginfo()
		#ͨ���˷����ж��Ƿ�������
		if (noDataFlag==1):
			#print "No data found current page=", i, ", QUIT"
			break

		#��ʱ��Ӧ�������ݣ����ǵõ�����������Ϊ0�����»�ȡ����
		if (count==0):
			print "Warnig: !!! Reload data in page=", i
			reloadUrl += 1
			if (reloadUrl>9):
				print "��ȡ���ݿ��ܲ��������������»�ȡ"
				break
			continue

		#���i��һ��������һҳ
		i += 1

	if addcsv==1:
		fcsv.close()
		if (totalline==0):
			os.remove(filecsv)

	if totalline>0:
		loginfo()
		ws = wb.create_sheet()
		write_statics(ws, fctime, dataObj, qdate)

	loginfo()
	filexlsx = prepath +filename+ '.xlsx'
	wb.save(filexlsx)
	if (totalline==0):
		os.remove(filexlsx)
		if bFindHist==1:
			handle_his_data(addcsv, prepath, hisUrl, code, qdate, stockInfo, sarr)
		else:
			print qdate+ " No Matched Record"
	else:
		print qdate+ " Saved OK"

def handle_his_data(addcsv, prepath, url, code, qdate, stockInfo, sarr):
	print "handle_his_data, url=",url
	if not os.path.isdir(prepath):
		os.makedirs(prepath)

	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)
	dataObjLen = len(dataObj)

	wb = Workbook()
	# grab the active worksheet
	ws = wb.active

	totalline = 0
	#���������ڲ�ͬ��ҳ�棬ͬʱ���ڣ������ظ�������Ҫ�����ظ����
	#��������ͬʱ�䣬��������ɽ�������Ҫ������
	lasttime = ''
	lastvol = 0
	pageFtime = ''
	bFtime = 0
	filename = code+ '_' + qdate
	if addcsv==1:
		filecsv = prepath + filename + '.csv'
		fcsv = open(filecsv, 'w')
		strline = '�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����'
		fcsv.write(strline)
		fcsv.write("\n")

	strline = u'�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����,���̼�,�ǵ���,ǰ�ռ�,���̼�,��߼�,��ͼ�,�ɽ���,�ɽ���'
	strObj = strline.split(u',')
	ws.append(strObj)
	dtlRe = re.compile(r'\D+(\d{2}:\d{2}:\d{2})\D+(\d+.\d{1,2})</td><td>(--|\+?\d+.\d+|-\d+.\d+)\D+(\d+)</td><td>([\d,]+)</td><th><h\d+>(����|����|������)\D')
	for i in range(1,1000):
		urlall = url + "&page=" +str(i)
		#print "%d, %s" %(i,urlall)

		req = urllib2.Request(urlall)
		res_data = urllib2.urlopen(req)

		flag = 0
		count = 0
		bFtime = 0
		line = res_data.readline()
		checkStr = '�ɽ�ʱ��'
		while line:
			#print line
			index = line.find(checkStr)
			if (index<0):
				line = res_data.readline()
				continue

			if flag==0:
				checkStr = '<th>'
				flag = 1
			else:
				key = dtlRe.match(line)
				if (key):
					curtime = key.group(1)
					price = key.group(2)
					fluctuate = key.group(3)
					curvol = int(key.group(4))
					intcurvol = int(key.group(4))
					amount = key.group(5)
					state = key.group(6)
					srange = ''
					
					#��ס��ǰҳ��һ����ʱ��
					if (bFtime==0):
						timeobj = re.search(curtime, pageFtime)
						if timeobj:
							break
						pageFtime = curtime
						bFtime = 1

					timeobj = re.search(curtime, lasttime)
					if (timeobj and intcurvol==lastvol):
						pass
					else:
						lasttime = curtime
						lastvol = intcurvol
						obj = amount.split(',')
						amount_n = ''.join(obj)
						intamount = int(amount_n)

						if cmp(state, '����')==0:
							handle_volumn(intcurvol, dataObj, 2)
						elif cmp(state, '����')==0:
							handle_volumn(intcurvol, dataObj, 1)
						elif cmp(state, '������')==0:
							ret = handle_middle_volumn(intcurvol, dataObj, curtime, fluctuate, key.group(3))
							if ret==1:
								state = '����'
							elif ret==2:
								state = '����'

						if addcsv==1:
							strline = curtime +","+ price +","+ srange +","+ fluctuate +","+ curvol +","+ amount_n +","+ state +"\n"
							fcsv.write(strline)

						totalline += 1
						row = totalline+1
						cell = 'A' + str(row)
						ws[cell] = curtime
						cell = 'B' + str(row)
						ws[cell] = float(price)
						cell = 'C' + str(row)
						ws[cell] = srange
						cell = 'D' + str(row)
						ws[cell] = fluctuate
						cell = 'E' + str(row)
						ws[cell] = curvol
						cell = 'F' + str(row)
						ws[cell] = intamount
						cell = 'G' + str(row)
						s1 = state.decode('gbk')
						ws[cell] = s1

						if row==2:
							ascid = 72
							number = len(stockInfo)
							for j in range(0,number):
								cell = chr(ascid+j) + str(row)
								ws[cell] = stockInfo[j]

					count += 1
				else:
					endObj = re.search(r'</td><td>', qdate)
					if (endObj):
						print "Error line:" + line
					else:
						break;
			line = res_data.readline()

		#���û���κ����ݵõ�����ʾ��ҳû�����ݣ��Ժ�ҳҲ��û�����ݣ��˳�ѭ��
		if (count==0):
			print "No data found i=", i, ", QUIT"
			break;

	if addcsv==1:
		fcsv.close()
		if (totalline==0):
			os.remove(filecsv)

	if (totalline>0):
		ws = wb.create_sheet()
		write_statics(ws, '', dataObj, qdate)

	filexlsx = prepath +filename+ '.xlsx'
	wb.save(filexlsx)
	if (totalline==0):
		print qdate +" No Matched Record!"
		os.remove(filexlsx)
	else:
		print qdate+ " Saved OK!"
