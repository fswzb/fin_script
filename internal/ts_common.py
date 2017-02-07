# -*- coding:gbk -*-
import sys
import re
import os
import string
import datetime
import urllib
import urllib2
from openpyxl import Workbook
from openpyxl.reader.excel  import  load_workbook
from common import *
import ctypes
import tushare as ts

def ts_handle_data(addcsv, prepath, bhist, url, code, qdate, sarr):
	todayUrl = "http://hq.sinajs.cn/list=" + code
	if bhist==0:
		url = url +"?symbol="+ code
	elif bhist==1:
		url = url +"?date="+ qdate +"&symbol="+ code
	#�������ڰ�����ʷ��¼����
	elif bhist==2:
		url = url +"?symbol="+ code
	else:
		print "Unknown flag:", bhist
		return -1
	#if Handle_Mid==0:
	#	print "Message: Ignore ������"

	if not os.path.isdir(prepath):
		os.makedirs(prepath)

	dataObj = []
	if cmp(sarr, '')==0:
		sarr = dftsarr
	volObj = sarr.split(',')
	arrlen = len(volObj)
	for i in range(0,arrlen):
		obj = fitItem(int(volObj[i]))
		dataObj.append(obj)

	wb = Workbook()
	# grab the active worksheet
	ws = wb.active

	totalline = 0
	filename = code+ '_' + qdate
	fctime = ''
	todayData = []
	todayDataLen = 0
	bGetToday = 0
	last_close = 0

	cur=datetime.datetime.now()
	if bhist==0:
		fctime = '%02d:%02d' %(cur.hour, cur.minute)
		filename = '%s_%02d-%02d' %(filename, cur.hour, cur.minute)
		bGetToday = 1
	if (bhist==2 and cur.hour>=15 and cur.minute>0):
		bGetToday = 1
	if (bGetToday==1):
		try:
			req = urllib2.Request(todayUrl)
			stockData = urllib2.urlopen(req, timeout=5).read()
		except:
			loginfo(1)
			print "URL timeout"
		else:
			stockObj = stockData.split(',')
			if len(stockObj)<10:
				print code, ": No trade data"
				return -1
			
			closePrice = float(stockObj[3])
			lastClsPrice = float(stockObj[2])
			last_close = lastClsPrice
			openPrice = float(stockObj[1])
			highPrice = float(stockObj[4])
			lowPrice = float(stockObj[5])
			exVolume = int(stockObj[8])/100
			exAmount = float(stockObj[9])
			f1 = '%02.02f'%( ((closePrice-lastClsPrice)/lastClsPrice)*100 )
			exFluc = float(f1)
			
			todayData.append(closePrice)
			todayData.append(exFluc)
			todayData.append(lastClsPrice)
			todayData.append(openPrice)
			todayData.append(highPrice)
			todayData.append(lowPrice)
			todayData.append(exVolume)
			todayData.append(exAmount)
			todayDataLen = len(todayData)

	if addcsv==1:
		filecsv = prepath + filename + '.csv'
		fcsv = open(filecsv, 'w')
		strline = '�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����'
		fcsv.write(strline)
		fcsv.write("\n")

	strline = u'�ɽ�ʱ��,�ɽ���,�ǵ���,�۸�䶯,�ɽ���,�ɽ���,����,���̼�,�ǵ���,ǰ�ռ�,���̼�,��߼�,��ͼ�,�ɽ���,�ɽ���'
	strObj = strline.split(u',')
	ws.append(strObj)

	stockInfo = []
	#ÿһҳ�����ݣ�����ҵ�ƥ������������Ϊ1�������ʱ��ҳ�������ݵ����ղ�����
	#countΪ0�����¼��س����ٴλ�ȡ��������������ݵ�ҳ�棬���countΪ0�Ͳ��ټ�����������
	matchDataFlag = 0
	excecount = 0
	savedTrasData = []
	savedTrasData2 = []
	largeTrasData = []
	i = 1

	curcode = code
	if len(code)==8:
		curcode = code[2:8]

	#��ʷ������������ͨ���˷������
	if bhist==1:
		df = ts.get_hist_data(curcode, start=qdate, end=qdate)
		if df is None or df.empty or len(df)!=1:
			print qdate, ": No data"
			return -1
		#print qdate, df
		for index,row in df.iterrows():
			last_close = float(row['close'])-float(row['price_change'])
			stockInfo.append(row['close'])
			stockInfo.append(row['p_change'])
			stockInfo.append(last_close)
			stockInfo.append(row['open'])
			stockInfo.append(row['high'])
			stockInfo.append(row['low'])
			stockInfo.append(row['volume'])
			stockInfo.append(row['turnover'])

	while excecount<=3:
		df = ts.get_tick_data(curcode, date=qdate)
		if df is None:
			excecount += 1
			continue;
		if df.size==18:
			excecount += 1
			continue;
		else:
			break;
		
	#����3�Σ������
	if df is None:
		print qdate, ": None Object"
		return -1
	if df.size==18:
		print qdate, ": Fail to get data"
		return -1

	for index,row in df.iterrows():
		curtime = row['time']
		curprice = row['price']
		range_per = ''
		if last_close!=0:
			range_val = ((float(curprice)-last_close) * 100) / last_close
			range_per = round(range_val, 2)
		fluctuate = row['change']
		curvol = int(row['volume'])
		volume = curvol
		amount = row['amount']
		state = row['type'].decode('utf8')
		#print state.decode('utf8')

		ret,hour,minute,second = parseTime(curtime)
		if (ret==-1):
			continue
		if (int(amount)==0 and not (hour==15 and minute==0)):
			continue

		bAddVolumn = 1
		if (hour==9 and minute==25) or (hour==15 and minute==0):
			bAddVolumn = 0

		stateStr = state
		st_buy = '����'.decode('gbk')
		st_sell = '����'.decode('gbk')
		st_mid = '������'.decode('gbk')
		if cmp(state, st_sell)==0:
			if bAddVolumn==1:
				handle_volumn(volume, dataObj, 2)
			stateStr = 'SELL����'.decode('gbk')
		elif cmp(state, st_buy)==0:
			if bAddVolumn==1:
				handle_volumn(volume, dataObj, 1)
		#Ŀǰ������û�д���
		elif cmp(state, st_mid)==0:
			if bAddVolumn==1:
				ret = handle_middle_volumn(volume, dataObj, curtime, fluctuate, 0)
			else:
				ret = 0
			if ret==1:
				stateStr = st_buy
			elif ret==2:
				stateStr = st_sell

		if addcsv==1:
			strline = curtime +","+ curprice +","+ range_per +","+ fluctuate +","+ curvol +","+ amount +","+ stateStr + "\n"
			fcsv.write(strline)

		totalline += 1
		row = totalline+1
		price = float(curprice)
		cell = 'A' + str(row)
		ws[cell] = curtime
		cell = 'B' + str(row)
		ws[cell] = price
		cell = 'C' + str(row)
		ws[cell] = range_per
		cell = 'D' + str(row)
		ftfluct = fluctuate
		if (fluctuate=='--'):
			ws[cell] = fluctuate
		else:
			ftfluct = float(fluctuate)
			ws[cell] = ftfluct
		cell = 'E' + str(row)
		ws[cell] = curvol
		cell = 'F' + str(row)
		ws[cell] = int(amount)
		cell = 'G' + str(row)
		s1 = stateStr
		ws[cell] = s1

		#�������������Sheetҳ�����
		if (row==2 and (bhist==0 or bhist==1 or (bhist==2 and cur.hour>=15)) and todayDataLen>0):
			ascid = 72
			for k in range(0, todayDataLen):
				cell = chr(ascid+k) + str(row)
				ws[cell] = todayData[k]

		#����ʼ�����ɽ����ݱ���
		bSaveFlag = 0
		if (totalline==1 or (totalline<4 and curvol>100)):
			bSaveFlag = 1
		elif (hour==9 and minute==30 and curvol>300) or (hour==9 and minute<30):
			bSaveFlag = 2
		if bSaveFlag==1 or bSaveFlag==2:
			rowData = []
			rowData.append(curtime)
			rowData.append(price)
			rowData.append(range_per)
			rowData.append(ftfluct)
			rowData.append(curvol)
			rowData.append(int(amount))
			rowData.append(s1)
		if bSaveFlag==1:
			savedTrasData.append(rowData)
		elif bSaveFlag==2:
			savedTrasData2.append(rowData)

		#���Ӵ󵥳ɽ���¼
		if (curvol>=Large_Volume):
			rowData = []
			rowData.append(curtime)
			rowData.append(price)
			rowData.append(range_per)
			rowData.append(ftfluct)
			rowData.append(curvol)
			rowData.append(int(amount))
			rowData.append(s1)
			largeTrasData.append(rowData)

		if (row==2 and bhist==1):
			ascid = 72
			number = len(stockInfo)
			for k in range(0,number):
				cell = chr(ascid+k) + str(row)
				ws[cell] = stockInfo[k]

	ws.auto_filter.ref = "A1:G1"

	if addcsv==1:
		fcsv.close()
		if (totalline==0):
			os.remove(filecsv)

	if totalline>0:
		startIdx = 0
		savedTrasLen = len(savedTrasData2)
		if savedTrasLen>0:
			if savedTrasLen>Tras_Count:
				startIdx = savedTrasLen-Tras_Count
			for j in range(startIdx, savedTrasLen):
				savedTrasData.append(savedTrasData2[j])
		ws = wb.create_sheet()
		write_statics(ws, fctime, dataObj, qdate, savedTrasData, largeTrasData)

	filexlsx = prepath +filename+ '.xlsx'
	if (os.path.exists(filexlsx) and bhist==0):
		j = 1
		while True:
			filexlsx = prepath + filename + '_' + str(j) + '.xlsx'
			j += 1
			if not os.path.exists(filexlsx):
				break;

	wb.save(filexlsx)
	return 0